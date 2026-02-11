import pandas as pd
from openpyxl import load_workbook

df = pd.read_excel("sales_data.xlsx", sheet_name="2025")

df["Total"] = df["Quantity"] * df["Price"]

df.to_excel("sales_summary.xlsx", index=False)

wb = load_workbook("sales_data.xlsx")
sheet = wb["2025"]

sheet["D1"] = "Total"

for row in range(2, sheet.max_row + 1):
    qty = sheet[f"B{row}"].value
    price = sheet[f"C{row}"].value
    sheet[f"D{row}"] = qty * price


wb.save("sales_summary_openpyxl.xlsx")
